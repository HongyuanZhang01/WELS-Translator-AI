"""
export_to_excel.py - Professional Excel report from evaluation results

USAGE:
  python -m pipeline.eval.export_to_excel                              # All results in eval/results/
  python -m pipeline.eval.export_to_excel --file eval/results/eval_pipeline_001.json
  python -m pipeline.eval.export_to_excel --dir pipeline/data/          # From a custom directory
  python -m pipeline.eval.export_to_excel --output my_report.xlsx       # Custom output path

OUTPUT: A professional Excel spreadsheet with 5 sheets:
  1. Dashboard     - Summary stats, averages, and a score distribution table
  2. Scores        - One row per chunk with all scores, color-coded
  3. Details       - Full explanations, critical errors, and suggestions
  4. Back-Translation - Meaning differences from the round-trip test
  5. Consistency   - Shows where dual-pass scores disagreed (reliability check)
"""

import os
import sys
import json
import glob
import argparse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from config import EVAL_OUTPUT_DIR, EVAL_WEIGHTS

# =========================================================================
# STYLE CONSTANTS
# =========================================================================
FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
SUBHEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="3D5A80")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1B2A4A")
SUBTITLE_FONT = Font(name=FONT_NAME, size=10, color="666666")
BODY_FONT = Font(name=FONT_NAME, size=10)
SCORE_FONT = Font(name=FONT_NAME, bold=True, size=11)

GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
OK_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

CRITICAL_FONT = Font(name=FONT_NAME, bold=True, color="CC0000", size=10)
GOOD_FONT = Font(name=FONT_NAME, bold=True, color="006100", size=10)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def score_fill(score):
    if isinstance(score, (int, float)):
        if score >= 4:
            return GOOD_FILL
        elif score >= 3:
            return OK_FILL
        return BAD_FILL
    return WHITE_FILL


def score_font(score):
    if isinstance(score, (int, float)):
        if score >= 4:
            return GOOD_FONT
        elif score < 3:
            return CRITICAL_FONT
    return SCORE_FONT


def style_cell(cell, font=None, fill=None, alignment=None, border=THIN_BORDER, number_format=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format


def write_header_row(ws, row, headers, widths=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        style_cell(cell, HEADER_FONT, HEADER_FILL, CENTER_WRAP)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


# =========================================================================
# DATA LOADING
# =========================================================================

def load_results(results_dir=None, single_file=None):
    if single_file:
        with open(single_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            data["_source_file"] = os.path.basename(single_file)
            return [data]

    results_dir = results_dir or os.path.abspath(EVAL_OUTPUT_DIR)
    files = sorted(glob.glob(os.path.join(results_dir, "eval_*.json")))

    # Also check for evaluations_*.json (pipeline output format)
    files += sorted(glob.glob(os.path.join(results_dir, "evaluations_*.json")))

    results = []
    seen_chunks = set()
    for fp in files:
        with open(fp, "r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            for item in data:
                key = (item.get("chunk_id"), item.get("source_label", ""))
                if key not in seen_chunks:
                    item["_source_file"] = os.path.basename(fp)
                    results.append(item)
                    seen_chunks.add(key)
        else:
            key = (data.get("chunk_id"), data.get("source_label", ""))
            if key not in seen_chunks:
                data["_source_file"] = os.path.basename(fp)
                results.append(data)
                seen_chunks.add(key)

    return results


# =========================================================================
# SHEET 1: DASHBOARD
# =========================================================================

def build_dashboard(wb, results):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "1B2A4A"

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value="WELS Translation Evaluation Report")
    style_cell(title_cell, TITLE_FONT, alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ts = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    sub_cell = ws.cell(row=2, column=1, value=f"Generated {ts} | {len(results)} chunk(s) evaluated")
    style_cell(sub_cell, SUBTITLE_FONT)

    # Summary stats
    dimensions = ["doctrinal_accuracy", "terminology_consistency", "clarity", "naturalness"]
    dim_labels = ["Doctrinal Accuracy", "Terminology", "Clarity", "Naturalness"]

    # Compute stats
    scores_by_dim = {d: [] for d in dimensions}
    weighted_scores = []
    for r in results:
        rubric = r.get("rubric_evaluation", {})
        if "error" in rubric:
            continue
        for d in dimensions:
            s = rubric.get(d, {}).get("score")
            if s is not None:
                scores_by_dim[d].append(s)
        ws_val = rubric.get("weighted_score")
        if ws_val is not None:
            weighted_scores.append(ws_val)

    row = 4
    summary_headers = ["Dimension", "Weight", "Average", "Min", "Max", "Count"]
    widths = [26, 10, 10, 10, 10, 10, 30]
    write_header_row(ws, row, summary_headers, widths)
    row += 1

    for dim, label in zip(dimensions, dim_labels):
        scores = scores_by_dim[dim]
        weight = EVAL_WEIGHTS.get(dim, 0)
        avg = sum(scores) / len(scores) if scores else 0
        mn = min(scores) if scores else 0
        mx = max(scores) if scores else 0

        ws.cell(row=row, column=1, value=label).font = Font(name=FONT_NAME, bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=weight).number_format = "0%"
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = CENTER

        avg_cell = ws.cell(row=row, column=3, value=round(avg, 2))
        style_cell(avg_cell, score_font(avg), score_fill(avg), CENTER, number_format="0.00")

        mn_cell = ws.cell(row=row, column=4, value=mn)
        style_cell(mn_cell, BODY_FONT, score_fill(mn), CENTER)

        mx_cell = ws.cell(row=row, column=5, value=mx)
        style_cell(mx_cell, BODY_FONT, score_fill(mx), CENTER)

        ws.cell(row=row, column=6, value=len(scores)).alignment = CENTER
        ws.cell(row=row, column=6).border = THIN_BORDER
        row += 1

    # Overall weighted
    row += 1
    ws.cell(row=row, column=1, value="WEIGHTED OVERALL").font = Font(name=FONT_NAME, bold=True, size=12, color="1B2A4A")
    if weighted_scores:
        avg_w = sum(weighted_scores) / len(weighted_scores)
        overall_cell = ws.cell(row=row, column=3, value=round(avg_w, 2))
        style_cell(overall_cell, Font(name=FONT_NAME, bold=True, size=14), score_fill(avg_w), CENTER, number_format="0.00")

    # Bar chart of averages
    if any(scores_by_dim[d] for d in dimensions):
        chart_row = row + 3
        ws.cell(row=chart_row, column=1, value="Dimension").font = SUBHEADER_FONT
        ws.cell(row=chart_row, column=2, value="Average Score").font = SUBHEADER_FONT
        for i, (dim, label) in enumerate(zip(dimensions, dim_labels)):
            scores = scores_by_dim[dim]
            avg = sum(scores) / len(scores) if scores else 0
            ws.cell(row=chart_row + 1 + i, column=1, value=label)
            ws.cell(row=chart_row + 1 + i, column=2, value=round(avg, 2))

        chart = BarChart()
        chart.type = "col"
        chart.title = "Average Scores by Dimension"
        chart.y_axis.title = "Score (1-5)"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 5
        chart.x_axis.title = None
        chart.style = 10
        chart.width = 20
        chart.height = 12

        cats = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + 4)
        vals = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row + 4)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4

        ws.add_chart(chart, f"D{chart_row}")

    # Back-translation summary
    bt_row = row + 3 + 6 + 2
    ws.cell(row=bt_row, column=1, value="Back-Translation Results").font = Font(name=FONT_NAME, bold=True, size=12, color="1B2A4A")
    bt_row += 1
    preserved = sum(1 for r in results
                    if r.get("back_translation", {}).get("comparison", {}).get("meaning_preserved") is True)
    not_preserved = sum(1 for r in results
                        if r.get("back_translation", {}).get("comparison", {}).get("meaning_preserved") is False)
    ws.cell(row=bt_row, column=1, value="Meaning preserved:")
    mp_cell = ws.cell(row=bt_row, column=2, value=preserved)
    style_cell(mp_cell, GOOD_FONT, GOOD_FILL, CENTER)
    bt_row += 1
    ws.cell(row=bt_row, column=1, value="Meaning NOT preserved:")
    mnp_cell = ws.cell(row=bt_row, column=2, value=not_preserved)
    if not_preserved > 0:
        style_cell(mnp_cell, CRITICAL_FONT, BAD_FILL, CENTER)
    else:
        style_cell(mnp_cell, GOOD_FONT, GOOD_FILL, CENTER)

    ws.freeze_panes = "A4"


# =========================================================================
# SHEET 2: SCORES
# =========================================================================

def build_scores_sheet(wb, results):
    ws = wb.create_sheet("Scores")
    ws.sheet_properties.tabColor = "3D5A80"

    headers = [
        "Chunk ID", "Source", "Language",
        "Doctrinal\nAccuracy", "Terminology", "Clarity", "Naturalness",
        "Weighted\nScore", "Meaning\nPreserved?", "# Diffs", "# Critical\nErrors"
    ]
    widths = [12, 16, 14, 14, 14, 12, 14, 14, 14, 10, 14]
    write_header_row(ws, 1, headers, widths)
    ws.row_dimensions[1].height = 32

    for row_idx, r in enumerate(results, 2):
        rubric = r.get("rubric_evaluation", {})
        bt = r.get("back_translation", {}).get("comparison", {})

        da = rubric.get("doctrinal_accuracy", {}).get("score", "")
        tc = rubric.get("terminology_consistency", {}).get("score", "")
        cl = rubric.get("clarity", {}).get("score", "")
        na = rubric.get("naturalness", {}).get("score", "")
        ws_score = rubric.get("weighted_score", "")

        meaning = bt.get("meaning_preserved", "")
        if isinstance(meaning, bool):
            meaning = "Yes" if meaning else "NO"

        diffs = len(bt.get("differences", []))
        errors = rubric.get("critical_errors", [])
        num_errors = len([e for e in errors if e]) if errors else 0

        values = [
            r.get("chunk_id", ""),
            r.get("source_label", r.get("pipeline_version", "")),
            r.get("target_language", ""),
            da, tc, cl, na, ws_score, meaning, diffs, num_errors
        ]

        is_alt_row = row_idx % 2 == 0
        row_fill = LIGHT_GRAY if is_alt_row else WHITE_FILL

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            cell.font = BODY_FONT

            if col in (4, 5, 6, 7) and isinstance(val, (int, float)):
                cell.fill = score_fill(val)
                cell.font = score_font(val)
            elif col == 8 and isinstance(val, (int, float)):
                cell.fill = score_fill(val)
                cell.font = score_font(val)
                cell.number_format = "0.00"
            elif col == 9:
                if val == "NO":
                    cell.fill = BAD_FILL
                    cell.font = CRITICAL_FONT
                elif val == "Yes":
                    cell.fill = GOOD_FILL
                    cell.font = GOOD_FONT
                else:
                    cell.fill = row_fill
            elif col == 11 and isinstance(val, int) and val > 0:
                cell.fill = BAD_FILL
                cell.font = CRITICAL_FONT
            else:
                cell.fill = row_fill

    ws.auto_filter.ref = f"A1:K{len(results) + 1}"
    ws.freeze_panes = "A2"


# =========================================================================
# SHEET 3: DETAILS
# =========================================================================

def build_details_sheet(wb, results):
    ws = wb.create_sheet("Details")
    ws.sheet_properties.tabColor = "98C1D9"

    headers = ["Chunk ID", "Source", "Dimension", "Score", "Explanation",
               "Critical Errors", "Suggestions"]
    widths = [12, 16, 24, 8, 70, 50, 50]
    write_header_row(ws, 1, headers, widths)

    dimensions = ["doctrinal_accuracy", "terminology_consistency", "clarity", "naturalness"]
    dim_labels = ["Doctrinal Accuracy", "Terminology Consistency", "Clarity", "Naturalness"]

    row = 2
    for r in results:
        rubric = r.get("rubric_evaluation", {})
        chunk_id = r.get("chunk_id", "")
        source = r.get("source_label", r.get("pipeline_version", ""))

        errors = rubric.get("critical_errors", [])
        errors_text = "\n".join([e for e in errors if e]) if errors else "None"
        suggestions = rubric.get("suggestions", [])
        suggestions_text = "\n".join([f"- {s}" for s in suggestions if s]) if suggestions else "None"

        for i, (dim, label) in enumerate(zip(dimensions, dim_labels)):
            dim_data = rubric.get(dim, {})
            score = dim_data.get("score", "")
            explanation = dim_data.get("explanation", "")

            ws.cell(row=row, column=1, value=chunk_id).border = THIN_BORDER
            ws.cell(row=row, column=1).font = BODY_FONT
            ws.cell(row=row, column=2, value=source).border = THIN_BORDER
            ws.cell(row=row, column=2).font = BODY_FONT
            ws.cell(row=row, column=3, value=label).border = THIN_BORDER
            ws.cell(row=row, column=3).font = Font(name=FONT_NAME, bold=True, size=10)

            score_cell = ws.cell(row=row, column=4, value=score)
            style_cell(score_cell, score_font(score), score_fill(score), CENTER)

            exp_cell = ws.cell(row=row, column=5, value=explanation)
            style_cell(exp_cell, BODY_FONT, alignment=LEFT_TOP)

            if i == 0:
                err_cell = ws.cell(row=row, column=6, value=errors_text)
                style_cell(err_cell, BODY_FONT if errors_text == "None" else CRITICAL_FONT, alignment=LEFT_TOP)
                sug_cell = ws.cell(row=row, column=7, value=suggestions_text)
                style_cell(sug_cell, BODY_FONT, alignment=LEFT_TOP)

            row += 1
        row += 1  # blank row between chunks

    ws.freeze_panes = "A2"


# =========================================================================
# SHEET 4: BACK-TRANSLATION
# =========================================================================

def build_backtranslation_sheet(wb, results):
    ws = wb.create_sheet("Back-Translation")
    ws.sheet_properties.tabColor = "EE6C4D"

    headers = ["Chunk ID", "Source", "Meaning\nPreserved?",
               "Original Phrase", "Back-Translated Phrase",
               "Severity", "Explanation", "Overall Assessment"]
    widths = [12, 16, 14, 40, 40, 12, 55, 55]
    write_header_row(ws, 1, headers, widths)
    ws.row_dimensions[1].height = 32

    severity_fills = {"critical": BAD_FILL, "moderate": OK_FILL, "minor": GOOD_FILL}
    severity_fonts = {"critical": CRITICAL_FONT, "moderate": BODY_FONT, "minor": BODY_FONT}

    row = 2
    for r in results:
        bt = r.get("back_translation", {}).get("comparison", {})
        chunk_id = r.get("chunk_id", "")
        source = r.get("source_label", r.get("pipeline_version", ""))

        meaning = bt.get("meaning_preserved", "")
        if isinstance(meaning, bool):
            meaning = "Yes" if meaning else "NO"

        diffs = bt.get("differences", [])
        assessment = bt.get("overall_assessment", "")

        if not diffs:
            ws.cell(row=row, column=1, value=chunk_id).border = THIN_BORDER
            ws.cell(row=row, column=2, value=source).border = THIN_BORDER
            mp_cell = ws.cell(row=row, column=3, value=meaning)
            style_cell(mp_cell, GOOD_FONT if meaning == "Yes" else CRITICAL_FONT,
                       GOOD_FILL if meaning == "Yes" else BAD_FILL, CENTER)
            ws.cell(row=row, column=4, value="No differences found").border = THIN_BORDER
            ws.cell(row=row, column=8, value=assessment).border = THIN_BORDER
            ws.cell(row=row, column=8).alignment = LEFT_TOP
            row += 1
        else:
            for i, diff in enumerate(diffs):
                ws.cell(row=row, column=1, value=chunk_id if i == 0 else "").border = THIN_BORDER
                ws.cell(row=row, column=2, value=source if i == 0 else "").border = THIN_BORDER

                if i == 0:
                    mp_cell = ws.cell(row=row, column=3, value=meaning)
                    style_cell(mp_cell, GOOD_FONT if meaning == "Yes" else CRITICAL_FONT,
                               GOOD_FILL if meaning == "Yes" else BAD_FILL, CENTER)
                else:
                    ws.cell(row=row, column=3, value="").border = THIN_BORDER

                ws.cell(row=row, column=4, value=diff.get("original_phrase", "")).border = THIN_BORDER
                ws.cell(row=row, column=4).alignment = LEFT_TOP
                ws.cell(row=row, column=5, value=diff.get("back_translated_phrase", "")).border = THIN_BORDER
                ws.cell(row=row, column=5).alignment = LEFT_TOP

                sev = diff.get("severity", "").lower()
                sev_cell = ws.cell(row=row, column=6, value=sev.upper() if sev else "")
                style_cell(sev_cell, severity_fonts.get(sev, BODY_FONT),
                           severity_fills.get(sev), CENTER)

                ws.cell(row=row, column=7, value=diff.get("explanation", "")).border = THIN_BORDER
                ws.cell(row=row, column=7).alignment = LEFT_TOP

                if i == 0:
                    ws.cell(row=row, column=8, value=assessment).border = THIN_BORDER
                    ws.cell(row=row, column=8).alignment = LEFT_TOP

                row += 1
        row += 1

    ws.freeze_panes = "A2"


# =========================================================================
# SHEET 5: CONSISTENCY
# =========================================================================

def build_consistency_sheet(wb, results):
    ws = wb.create_sheet("Consistency")
    ws.sheet_properties.tabColor = "293241"

    headers = ["Chunk ID", "Source", "Dimension",
               "Pass A", "Pass B", "Pass C\n(tiebreaker)", "Final Score",
               "Agreed?", "Tiebreaker\nUsed?"]
    widths = [12, 16, 24, 10, 10, 14, 12, 10, 14]
    write_header_row(ws, 1, headers, widths)
    ws.row_dimensions[1].height = 32

    dimensions = ["doctrinal_accuracy", "terminology_consistency", "clarity", "naturalness"]
    dim_labels = ["Doctrinal Accuracy", "Terminology", "Clarity", "Naturalness"]

    row = 2
    for r in results:
        rubric = r.get("rubric_evaluation", {})
        consistency = rubric.get("consistency_report", {})
        chunk_id = r.get("chunk_id", "")
        source = r.get("source_label", r.get("pipeline_version", ""))

        if not consistency:
            continue

        for dim, label in zip(dimensions, dim_labels):
            info = consistency.get(dim, {})
            all_scores = info.get("all_scores", [])
            agreed = info.get("agreed", True)
            final = info.get("final_score", "")
            tiebreaker = info.get("tiebreaker_used", False)

            ws.cell(row=row, column=1, value=chunk_id).border = THIN_BORDER
            ws.cell(row=row, column=2, value=source).border = THIN_BORDER
            ws.cell(row=row, column=3, value=label).border = THIN_BORDER
            ws.cell(row=row, column=3).font = Font(name=FONT_NAME, bold=True, size=10)

            # Pass A
            if len(all_scores) > 0:
                ws.cell(row=row, column=4, value=all_scores[0]).border = THIN_BORDER
                ws.cell(row=row, column=4).alignment = CENTER
            # Pass B
            if len(all_scores) > 1:
                ws.cell(row=row, column=5, value=all_scores[1]).border = THIN_BORDER
                ws.cell(row=row, column=5).alignment = CENTER
            # Pass C
            if len(all_scores) > 2:
                ws.cell(row=row, column=6, value=all_scores[2]).border = THIN_BORDER
                ws.cell(row=row, column=6).alignment = CENTER

            final_cell = ws.cell(row=row, column=7, value=final)
            style_cell(final_cell, score_font(final), score_fill(final), CENTER)

            agreed_cell = ws.cell(row=row, column=8, value="Yes" if agreed else "NO")
            if agreed:
                style_cell(agreed_cell, GOOD_FONT, GOOD_FILL, CENTER)
            else:
                style_cell(agreed_cell, CRITICAL_FONT, BAD_FILL, CENTER)

            tb_cell = ws.cell(row=row, column=9, value="Yes" if tiebreaker else "-")
            style_cell(tb_cell, BODY_FONT, OK_FILL if tiebreaker else WHITE_FILL, CENTER)

            row += 1
        row += 1  # gap between chunks

    ws.freeze_panes = "A2"


# =========================================================================
# MAIN
# =========================================================================

def main():
    parser = argparse.ArgumentParser(description="Export evaluation results to professional Excel report")
    parser.add_argument("--file", type=str, help="Single JSON result file to export")
    parser.add_argument("--dir", type=str, help="Directory containing evaluation JSON files")
    parser.add_argument("--output", type=str, help="Output Excel file path")
    args = parser.parse_args()

    results_dir = args.dir or None
    results = load_results(results_dir=results_dir, single_file=args.file)
    if not results:
        print("No evaluation results found. Run the evaluator first.")
        return

    print(f"Loaded {len(results)} evaluation result(s)")

    wb = Workbook()
    build_dashboard(wb, results)
    build_scores_sheet(wb, results)
    build_details_sheet(wb, results)
    build_backtranslation_sheet(wb, results)
    build_consistency_sheet(wb, results)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = args.output or os.path.join(
        os.path.abspath(EVAL_OUTPUT_DIR), f"evaluation_report_{timestamp}.xlsx"
    )
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Report saved: {output_path}")


if __name__ == "__main__":
    main()
