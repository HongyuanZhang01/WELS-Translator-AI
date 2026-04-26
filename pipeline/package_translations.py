"""
package_translations.py - Build the team deliverable for a single translation job

WHAT THIS IS FOR
================
Once a document has been translated and evaluated by the pipeline, the
team needs polished artifacts they can read and review. This script takes
the raw pipeline outputs (chunks, translations, evaluations) for ONE job
and produces four deliverable files in a delivery folder.

It handles any translation job that follows the pipeline's standard
folder layout — past jobs, the ones running right now, and any future
document the team runs through the pipeline. It does not hardcode any
document names or language pairs.

WHAT IT PRODUCES (four files per job)
=====================================
    1. <job>_chunk_by_chunk_comparison.txt
         Every chunk shown side-by-side: source text, target translation,
         article label, and the four rubric scores. Chunks that need
         human review get a *** FLAGGED *** banner.

    2. <job>_<target>_full.txt
         The complete target-language translation only, cleanly
         concatenated. No English, no scores, nothing the reviewer
         doesn't want to see.

    3. <job>_chunks_needing_human_review.txt
         A focused list of just the flagged chunks, with the scores,
         the AI's explanation of what might be off, source text, and
         translation. This is where human reviewers should spend their
         attention.

    4. <job>_evaluation_summary.xlsx
         Excel file with a per-chunk scores table and averages. Good for
         a quick quantitative overview. Can be skipped with --no-excel.

FLAGGING RULE
=============
A chunk is flagged for human review if:
    doctrinal_accuracy.score < 5.0  OR  terminology_consistency.score < 5.0
Clarity and naturalness are reported but don't drive flagging.

USAGE
=====
    # See which jobs the packager can handle right now
    python -m pipeline.package_translations --list

    # Package one job
    python -m pipeline.package_translations --job wauwatosa_spanish

    # Interactive mode: pick one job from a numbered menu
    python -m pipeline.package_translations

    # Point at a different batch folder or output folder
    python -m pipeline.package_translations --job otto_hungarian \
        --batch-dir pipeline/data/batch_april2026 \
        --output translation_deliveries

    # Skip the Excel file (no openpyxl dependency)
    python -m pipeline.package_translations --job otto_spanish --no-excel
"""

import os
import sys
import glob
import json
import argparse
import re
from datetime import datetime


# Project conventions
DEFAULT_BATCH_DIR = "pipeline/data/batch_april2026"
DEFAULT_OUTPUT_PARENT = "translation_deliveries"

# Flag threshold — see module docstring
DOCTRINAL_THRESHOLD = 5.0
TERMINOLOGY_THRESHOLD = 5.0

TIMESTAMP_IN_NAME = re.compile(r"(\d{8}_\d{6})")

# File-name suffixes that mean "this is a post-processed canonical version,
# prefer it over raw pipeline-run files." Anything matching one of these
# tokens in its filename is treated the same way: canonical.
#   _MERGED  — produced by merge_eval_outputs.py, stitches partial+resume runs
#   _FIXED   — produced by hand to patch specific broken chunks in a run
CANONICAL_SUFFIX_TOKENS = ("_MERGED", "_FIXED")


def is_canonical_name(path):
    """Return True if this filename has any of the canonical suffix tokens
    in it, meaning a human or tool has already post-processed it."""
    base = os.path.basename(path)
    return any(tok in base for tok in CANONICAL_SUFFIX_TOKENS)


# =====================================================================
# JOB DISCOVERY AND FILE SELECTION
# =====================================================================

def list_job_folders(batch_dir):
    """Return a sorted list of subfolders under batch_dir that look like
    job folders (they contain at least one chunks_*.json)."""
    if not os.path.isdir(batch_dir):
        return []
    result = []
    for name in sorted(os.listdir(batch_dir)):
        path = os.path.join(batch_dir, name)
        if not os.path.isdir(path):
            continue
        if glob.glob(os.path.join(path, "chunks_*.json")):
            result.append(name)
    return result


def assess_job_readiness(job_folder):
    """Look at the files in a job folder and return a short status string
    describing whether it's ready to package. Used by --list."""
    evals = [f for f in glob.glob(os.path.join(job_folder, "evaluations_*.json"))]
    etrans = [f for f in glob.glob(os.path.join(job_folder, "evaluated_translations_*.json"))]
    canonical_evals = [f for f in evals if is_canonical_name(f)]
    canonical_etrans = [f for f in etrans if is_canonical_name(f)]
    raw_evals = [f for f in evals if not is_canonical_name(f)]
    raw_etrans = [f for f in etrans if not is_canonical_name(f)]

    if canonical_evals and canonical_etrans:
        return "READY (canonical files present)"
    if canonical_evals and len(raw_etrans) == 1:
        return "READY (canonical eval + single translations pair)"
    if len(raw_evals) == 1 and len(raw_etrans) == 1:
        return "READY (single clean run)"
    if len(raw_evals) == 0:
        return "NOT READY — no eval output yet"
    if len(raw_evals) > 1:
        return f"NEEDS MERGE — {len(raw_evals)} eval runs in folder"
    return "PROBLEM — eval/translation file count mismatch"


def select_eval_pair(job_folder):
    """Pick which evaluations file and evaluated_translations file to use.

    Priority:
        1. If a canonical pair (files with _MERGED or _FIXED in name) is
           present, use the newest canonical pair.
        2. Otherwise, require exactly one raw eval file and exactly one
           matching raw evaluated_translations file.
        3. Mixed case: if canonical evaluations exist but only one raw
           evaluated_translations file exists, pair them (this is the
           common 'fixed scores, untouched translations' pattern from
           wauwatosa_hungarian).

    Returns (evaluations_path, evaluated_translations_path) or raises
    SystemExit with a clear message explaining what's wrong."""
    all_evals = sorted(glob.glob(os.path.join(job_folder, "evaluations_*.json")))
    all_etrans = sorted(glob.glob(os.path.join(job_folder, "evaluated_translations_*.json")))

    canonical_evals = [f for f in all_evals if is_canonical_name(f)]
    canonical_etrans = [f for f in all_etrans if is_canonical_name(f)]
    raw_evals = [f for f in all_evals if not is_canonical_name(f)]
    raw_etrans = [f for f in all_etrans if not is_canonical_name(f)]

    # Case 1: full canonical pair.
    if canonical_evals and canonical_etrans:
        canonical_evals.sort(key=lambda p: os.path.getmtime(p))
        canonical_etrans.sort(key=lambda p: os.path.getmtime(p))
        return canonical_evals[-1], canonical_etrans[-1]

    # Case 3: canonical evaluations but plain evaluated_translations.
    if canonical_evals and not canonical_etrans:
        if len(raw_etrans) == 1:
            canonical_evals.sort(key=lambda p: os.path.getmtime(p))
            return canonical_evals[-1], raw_etrans[0]
        raise SystemExit(
            f"ERROR: found a canonical evaluations file but {len(raw_etrans)} "
            f"evaluated_translations files in {job_folder}. I don't know "
            f"which one pairs with the canonical evaluations. Move extras "
            f"into an archive folder, then re-run."
        )

    # Case 4: canonical evaluated_translations but plain evaluations — rare.
    if canonical_etrans and not canonical_evals:
        raise SystemExit(
            f"ERROR: found a canonical evaluated_translations file but no "
            f"canonical evaluations file in {job_folder}. Something is "
            f"inconsistent. Check the folder contents."
        )

    # Case 2: raw-only, require exactly one of each.
    if len(raw_evals) == 0:
        raise SystemExit(
            f"ERROR: no evaluations_*.json files found in {job_folder}. "
            f"Has the eval pipeline been run on this job yet?"
        )
    if len(raw_evals) > 1:
        raise SystemExit(
            f"ERROR: {len(raw_evals)} evaluations_*.json files found in\n  {job_folder}\n"
            f"This usually means you had a partial run + a resume run.\n"
            f"Merge them first with:\n"
            f"  python -m pipeline.merge_eval_outputs {job_folder}"
        )
    if len(raw_etrans) != 1:
        raise SystemExit(
            f"ERROR: {len(raw_etrans)} evaluated_translations_*.json files found, "
            f"expected 1. Folder: {job_folder}"
        )

    # Confirm the single raw pair shares a timestamp — otherwise something
    # strange happened in the folder.
    eval_ts = TIMESTAMP_IN_NAME.search(os.path.basename(raw_evals[0]))
    etrans_ts = TIMESTAMP_IN_NAME.search(os.path.basename(raw_etrans[0]))
    if eval_ts and etrans_ts and eval_ts.group(1) != etrans_ts.group(1):
        print(
            f"WARNING: evaluations and evaluated_translations have different "
            f"timestamps ({eval_ts.group(1)} vs {etrans_ts.group(1)}). "
            f"Proceeding anyway — this may produce mismatched output."
        )
    return raw_evals[0], raw_etrans[0]


def load_json(path, description):
    """Load a JSON file with a helpful error if it's malformed or truncated."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        raise SystemExit(
            f"ERROR: {description} at {path} is not valid JSON.\n"
            f"  Parse error: {e}\n"
            f"  If this file came from a crashed pipeline run, it may be "
            f"truncated mid-write. Look for an earlier complete version, "
            f"or re-run the eval pipeline."
        )


def load_bible_lookup_review(job_folder):
    """Load a bible_lookup_review_*.json file from the job folder, if one
    exists.

    These files are produced by run_pipeline.py's quote_handler when the
    Anthropic API's content filter refuses to return canonical Bible verse
    text for a reference. The translator writes its own inline rendering
    of the verse, but the failure gets recorded in this sidecar file so a
    human can manually verify the rendering against the published
    target-language Bible before delivery.

    Returns:
        None if no file is present.
        A dict {"data": <parsed json>, "source_file": <basename>} if a file
        was found and parsed successfully.

    If multiple bible_lookup_review files exist in the folder (e.g. from
    several runs), the NEWEST one by modification time is used. This is
    intentionally lenient — unlike eval/translation files, the bible-review
    file is additive metadata and a stale copy from a prior run shouldn't
    block packaging. We log the choice so it's never silent.
    """
    matches = sorted(
        glob.glob(os.path.join(job_folder, "bible_lookup_review_*.json")),
        key=lambda p: os.path.getmtime(p),
    )
    if not matches:
        return None
    newest = matches[-1]
    if len(matches) > 1:
        print(
            f"  NOTE: found {len(matches)} bible_lookup_review_*.json files; "
            f"using newest: {os.path.basename(newest)}"
        )
    try:
        with open(newest, "r", encoding="utf-8") as f:
            return {"data": json.load(f), "source_file": os.path.basename(newest)}
    except (json.JSONDecodeError, OSError) as e:
        print(
            f"  WARNING: could not read {os.path.basename(newest)}: {e}\n"
            f"  Proceeding without Bible verse review file."
        )
        return None


# =====================================================================
# RUBRIC ACCESS (brittle-safe)
# =====================================================================

def safe_score(rubric, key):
    """Pull a score out of the rubric evaluation dict with a clear error if
    the schema has changed since this script was written."""
    try:
        return float(rubric[key]["score"])
    except (KeyError, TypeError, ValueError) as e:
        raise SystemExit(
            f"ERROR: couldn't read rubric field '{key}.score' from an evaluation.\n"
            f"  Underlying error: {e}\n"
            f"  The pipeline's evaluation schema may have changed. Expected:\n"
            f"    rubric_evaluation.{key}.score (a number)\n"
            f"  If the schema changed, update safe_score() in this script."
        )


def safe_explanation(rubric, key):
    """Pull an explanation string out of the rubric with a safe fallback."""
    try:
        return str(rubric[key].get("explanation", "(no explanation)"))
    except (KeyError, TypeError):
        return "(no explanation)"


def extract_chunk_scores(eval_item):
    """Given one entry from evaluations_*.json, return a dict of the scores
    we care about plus a flagged flag."""
    rubric = eval_item.get("rubric_evaluation")
    if not isinstance(rubric, dict):
        raise SystemExit(
            f"ERROR: chunk {eval_item.get('chunk_id', '?')} has no "
            f"rubric_evaluation. The evaluations file may be incomplete."
        )
    scores = {
        "doctrinal": safe_score(rubric, "doctrinal_accuracy"),
        "terminology": safe_score(rubric, "terminology_consistency"),
        "clarity": safe_score(rubric, "clarity"),
        "naturalness": safe_score(rubric, "naturalness"),
    }
    # weighted_score may not exist on all schemas; fall back to average.
    try:
        scores["weighted"] = float(rubric.get("weighted_score", 0)) or sum(
            scores[k] for k in ("doctrinal", "terminology", "clarity", "naturalness")
        ) / 4.0
    except (TypeError, ValueError):
        scores["weighted"] = sum(
            scores[k] for k in ("doctrinal", "terminology", "clarity", "naturalness")
        ) / 4.0
    scores["flagged"] = (
        scores["doctrinal"] < DOCTRINAL_THRESHOLD
        or scores["terminology"] < TERMINOLOGY_THRESHOLD
    )
    return scores


# =====================================================================
# DELIVERABLE WRITERS
# =====================================================================

def write_chunk_by_chunk(output_path, job_name, target_language, source_language,
                         rows, generated_at):
    """Build the big side-by-side comparison file."""
    flagged_count = sum(1 for r in rows if r["scores"]["flagged"])
    total = len(rows)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("=" * 72 + "\n")
        f.write(f"{job_name.upper()} — Chunk-by-chunk review document\n")
        f.write(f"Target language: {target_language}\n")
        f.write(f"Source language: {source_language}\n")
        f.write(f"Total chunks: {total}\n")
        f.write(f"Chunks flagged for human review: {flagged_count}\n")
        f.write(f"Generated: {generated_at}\n")
        f.write("=" * 72 + "\n\n")

        f.write("HOW TO READ THIS FILE\n")
        f.write("-" * 72 + "\n")
        f.write(
            "Each chunk is shown with its source text, the translation, and\n"
            "the AI rubric scores (all out of 5). Chunks flagged with\n"
            "*** FLAGGED *** scored below 5 on either doctrinal accuracy or\n"
            "terminology consistency — those are where human reviewers should\n"
            "focus.\n\n"
        )

        for row in rows:
            scores = row["scores"]
            flag_banner = ""
            if scores["flagged"]:
                reasons = []
                if scores["doctrinal"] < DOCTRINAL_THRESHOLD:
                    reasons.append(f"doctrinal {scores['doctrinal']:.1f}")
                if scores["terminology"] < TERMINOLOGY_THRESHOLD:
                    reasons.append(f"terminology {scores['terminology']:.1f}")
                flag_banner = (
                    "\n*** FLAGGED — HUMAN REVIEW NEEDED "
                    f"({', '.join(reasons)}) ***\n"
                )
            f.write("=" * 72 + "\n")
            f.write(f"CHUNK {row['chunk_id']}")
            if row.get("position"):
                f.write(f"  (position {row['position']})")
            f.write("\n")
            if row.get("article"):
                f.write(f"Article: {row['article']}\n")
            f.write(
                f"Scores — Doctrinal: {scores['doctrinal']:.1f}  "
                f"Terminology: {scores['terminology']:.1f}  "
                f"Clarity: {scores['clarity']:.1f}  "
                f"Naturalness: {scores['naturalness']:.1f}  "
                f"Weighted: {scores['weighted']:.2f}\n"
            )
            if flag_banner:
                f.write(flag_banner)
            f.write("=" * 72 + "\n\n")

            f.write(f"--- SOURCE ({source_language}) ---\n")
            f.write(row["source_text"].rstrip() + "\n\n")
            f.write(f"--- TRANSLATION ({target_language}) ---\n")
            f.write(row["translated_text"].rstrip() + "\n\n\n")


def write_target_only(output_path, job_name, target_language, source_language,
                      rows, generated_at):
    """Build the clean translation-only file that's nice to read start to finish."""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(f"{job_name.upper()} — Complete {target_language} translation\n")
        f.write(f"Source language: {source_language}\n")
        f.write(f"Generated: {generated_at}\n")
        f.write("=" * 72 + "\n\n")
        for row in rows:
            f.write(row["translated_text"].rstrip() + "\n\n")


def write_review_list(output_path, job_name, target_language, rows, generated_at):
    """Build the focused list of chunks that need human review."""
    flagged = [r for r in rows if r["scores"]["flagged"]]
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("=" * 72 + "\n")
        f.write(f"{job_name.upper()} — Chunks needing human review\n")
        f.write(f"Target language: {target_language}\n")
        f.write(f"Total chunks: {len(rows)}\n")
        f.write(f"Flagged (doctrinal < 5 OR terminology < 5): {len(flagged)}\n")
        f.write(f"Generated: {generated_at}\n")
        f.write("=" * 72 + "\n\n")

        if not flagged:
            f.write("No chunks were flagged. Every chunk scored a perfect 5 on\n")
            f.write("both doctrinal accuracy and terminology consistency.\n")
            return

        for row in flagged:
            scores = row["scores"]
            f.write(f"## CHUNK {row['chunk_id']} — FLAGGED\n")
            if row.get("article"):
                f.write(f"Article: {row['article']}\n")
            f.write(
                f"Scores — Doctrinal: {scores['doctrinal']:.1f}  "
                f"Terminology: {scores['terminology']:.1f}  "
                f"Clarity: {scores['clarity']:.1f}  "
                f"Naturalness: {scores['naturalness']:.1f}\n\n"
            )
            f.write("Doctrinal accuracy explanation:\n")
            f.write(row["doctrinal_explanation"].strip() + "\n\n")
            f.write("Terminology consistency explanation:\n")
            f.write(row["terminology_explanation"].strip() + "\n\n")

            suggestions = row.get("suggestions") or []
            if suggestions:
                f.write("AI suggestions:\n")
                for s in suggestions:
                    f.write(f"  - {s.strip()}\n")
                f.write("\n")

            f.write("Source text:\n")
            f.write(row["source_text"].strip() + "\n\n")
            f.write("Translation:\n")
            f.write(row["translated_text"].strip() + "\n\n")
            f.write("-" * 72 + "\n\n")


def write_bible_verse_review_file(output_path, job_name, target_language,
                                  bible_review, rows, generated_at):
    """Build a plain-text file listing the Bible verse references the
    Anthropic API's content filter refused to return during translation.

    The translator still produced an inline rendering of each verse so the
    chunk could be translated — but a human reviewer needs to compare that
    inline rendering against the canonical published target-language Bible
    before the translation can be delivered.

    The `bible_review` argument is the dict returned by
    load_bible_lookup_review(). This function should only be called when
    that returns non-None.

    `rows` is the list of row dicts produced by build_rows(); we use it to
    pull out the source_text and translated_text for any chunk that had a
    Bible lookup failure, so reviewers have context without flipping back
    and forth between files.
    """
    data = bible_review["data"]
    failures = data.get("failures") or []
    rows_by_id = {r["chunk_id"]: r for r in rows}

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("=" * 72 + "\n")
        f.write(f"{job_name.upper()} — Bible verses needing human verification\n")
        f.write(f"Target language: {target_language}\n")
        f.write(f"Total failures: {len(failures)}\n")
        f.write(f"Source review file: {bible_review['source_file']}\n")
        f.write(f"Generated: {generated_at}\n")
        f.write("=" * 72 + "\n\n")

        f.write("WHAT THIS FILE IS\n")
        f.write("-" * 72 + "\n")
        f.write(
            "During translation, the AI tried to look up the canonical text\n"
            "of each Bible verse quoted in the source document so it could\n"
            "drop in the official published translation. For the references\n"
            "listed below, the lookup failed (the API's content filter\n"
            "refused to return the verse text). The translator fell back to\n"
            "producing its OWN rendering of the verse inside the translated\n"
            "chunk.\n\n"
            "A human reviewer must open the published target-language Bible,\n"
            "find each reference below, and compare it against the verse\n"
            "text that appears in the corresponding chunk of the delivered\n"
            "translation. If the translator's rendering is wrong, replace it\n"
            "with the published text before sending the translation out.\n\n"
        )

        if not failures:
            f.write("No Bible lookup failures recorded. Nothing to review here.\n")
            return

        for i, failure in enumerate(failures, start=1):
            chunk_id = failure.get("chunk_id", "?")
            references = failure.get("references") or failure.get("reference") or []
            if isinstance(references, str):
                references = [references]
            reason = failure.get("reason") or failure.get("error") or ""

            f.write("=" * 72 + "\n")
            f.write(f"FAILURE {i} of {len(failures)} — CHUNK {chunk_id}\n")
            f.write("=" * 72 + "\n\n")

            if references:
                f.write("References to verify:\n")
                for ref in references:
                    f.write(f"  - {ref}\n")
                f.write("\n")

            if reason:
                f.write(f"API reason: {reason}\n\n")

            row = rows_by_id.get(chunk_id)
            if row:
                if row.get("article"):
                    f.write(f"Article: {row['article']}\n\n")
                f.write("Source text (contains the verse reference):\n")
                f.write(row["source_text"].strip() + "\n\n")
                f.write("Translator's rendering (verify against published Bible):\n")
                f.write(row["translated_text"].strip() + "\n\n")
            else:
                f.write(
                    f"(Could not find chunk {chunk_id} in the evaluated\n"
                    f"translations file — review the verse references manually.)\n\n"
                )
            f.write("-" * 72 + "\n\n")


def write_excel_summary(output_path, job_name, target_language, source_language,
                        rows, generated_at):
    """Build the Excel summary file. Returns True on success, or prints a
    warning and returns False if openpyxl isn't installed."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print(
            "  WARNING: openpyxl not installed, skipping Excel summary.\n"
            "  To enable: pip install openpyxl"
        )
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"

    bold = Font(bold=True)
    flag_fill = PatternFill("solid", fgColor="FFE0E0")  # light red
    header_fill = PatternFill("solid", fgColor="D9E1F2")  # light blue

    # Header info rows
    ws.append([f"{job_name.upper()} — Evaluation summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Target language: {target_language}"])
    ws.append([f"Source language: {source_language}"])
    ws.append([f"Generated: {generated_at}"])
    ws.append([])  # spacer

    header_row_index = ws.max_row + 1
    headers = [
        "Chunk ID", "Position", "Article",
        "Doctrinal", "Terminology", "Clarity", "Naturalness", "Weighted",
        "Flagged?",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_index, column=col_idx)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in rows:
        s = row["scores"]
        data_row = [
            row["chunk_id"],
            row.get("position", ""),
            (row.get("article") or "")[:120],
            round(s["doctrinal"], 2),
            round(s["terminology"], 2),
            round(s["clarity"], 2),
            round(s["naturalness"], 2),
            round(s["weighted"], 2),
            "YES" if s["flagged"] else "",
        ]
        ws.append(data_row)
        if s["flagged"]:
            for col_idx in range(1, len(data_row) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = flag_fill

    # Averages row
    n = len(rows)
    if n:
        avg_row = [
            "AVERAGES", "", "",
            round(sum(r["scores"]["doctrinal"] for r in rows) / n, 2),
            round(sum(r["scores"]["terminology"] for r in rows) / n, 2),
            round(sum(r["scores"]["clarity"] for r in rows) / n, 2),
            round(sum(r["scores"]["naturalness"] for r in rows) / n, 2),
            round(sum(r["scores"]["weighted"] for r in rows) / n, 2),
            f"{sum(1 for r in rows if r['scores']['flagged'])}/{n}",
        ]
        ws.append([])
        ws.append(avg_row)
        for col_idx in range(1, len(avg_row) + 1):
            ws.cell(row=ws.max_row, column=col_idx).font = bold

    # Column widths
    widths = [10, 10, 48, 11, 13, 10, 13, 11, 10]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = w

    wb.save(output_path)
    return True


# =====================================================================
# MAIN PACKAGE FLOW
# =====================================================================

def build_rows(evaluations, evaluated_translations):
    """Join the evaluation and evaluated_translations lists by chunk_id and
    return a list of row dicts ready for the writer functions."""
    etrans_by_id = {c["chunk_id"]: c for c in evaluated_translations}
    missing = [
        e["chunk_id"] for e in evaluations
        if e["chunk_id"] not in etrans_by_id
    ]
    if missing:
        raise SystemExit(
            f"ERROR: evaluations file has chunks {missing} that are not in "
            f"the evaluated_translations file. Files appear to be out of sync."
        )

    rows = []
    for eval_item in evaluations:
        cid = eval_item["chunk_id"]
        etrans_item = etrans_by_id[cid]
        rubric = eval_item["rubric_evaluation"]
        row = {
            "chunk_id": cid,
            "position": etrans_item.get("position", ""),
            "article": etrans_item.get("article", ""),
            "source_text": etrans_item.get("source_text", "(missing)"),
            "translated_text": etrans_item.get("translated_text", "(missing)"),
            "scores": extract_chunk_scores(eval_item),
            "doctrinal_explanation": safe_explanation(rubric, "doctrinal_accuracy"),
            "terminology_explanation": safe_explanation(rubric, "terminology_consistency"),
            "suggestions": rubric.get("suggestions") or [],
        }
        rows.append(row)

    # Sort by chunk_id (lex order works because they're zero-padded: 001, 002...).
    rows.sort(key=lambda r: r["chunk_id"])
    return rows


def package_job(job_name, job_folder, output_parent, make_excel):
    """Full pipeline for packaging one job. Returns the path to the
    created delivery folder."""
    print(f"\n{'=' * 60}")
    print(f"PACKAGING: {job_name}")
    print(f"Source folder: {job_folder}")
    print("=" * 60)

    eval_path, etrans_path = select_eval_pair(job_folder)
    print(f"  evaluations:            {os.path.basename(eval_path)}")
    print(f"  evaluated_translations: {os.path.basename(etrans_path)}")

    # Optional: Bible-verse review sidecar. Not required for packaging.
    bible_review = load_bible_lookup_review(job_folder)
    if bible_review:
        n_failures = len(bible_review["data"].get("failures") or [])
        print(f"  bible_lookup_review:    {bible_review['source_file']} "
              f"({n_failures} failure(s))")

    evaluations = load_json(eval_path, "evaluations file")
    etrans = load_json(etrans_path, "evaluated_translations file")
    if not isinstance(evaluations, list) or not isinstance(etrans, list):
        raise SystemExit("ERROR: eval files must contain JSON lists at the top level.")
    if len(evaluations) != len(etrans):
        print(
            f"  WARNING: evaluations has {len(evaluations)} chunks but "
            f"evaluated_translations has {len(etrans)}. Proceeding with the "
            f"intersection by chunk_id."
        )

    # Language info — read from the first evaluated_translations item since
    # those carry both source_language and target_language per chunk.
    if not etrans:
        raise SystemExit("ERROR: evaluated_translations file is empty.")
    target_language = etrans[0].get("target_language", "Unknown")
    source_language = etrans[0].get("source_language", "Unknown")
    print(f"  target language:        {target_language}")
    print(f"  source language:        {source_language}")

    rows = build_rows(evaluations, etrans)
    flagged_count = sum(1 for r in rows if r["scores"]["flagged"])
    print(f"  chunks in package:      {len(rows)}")
    print(f"  flagged for review:     {flagged_count}")

    # Create the delivery folder.
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    delivery_name = f"{job_name}_delivery_{timestamp}"
    delivery_folder = os.path.join(output_parent, delivery_name)
    os.makedirs(delivery_folder, exist_ok=True)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    target_lower = target_language.lower()

    # 1. Chunk-by-chunk comparison
    cbc_path = os.path.join(delivery_folder, f"{job_name}_chunk_by_chunk_comparison.txt")
    write_chunk_by_chunk(cbc_path, job_name, target_language, source_language,
                         rows, generated_at)
    print(f"  wrote {os.path.basename(cbc_path)}")

    # 2. Full target-only text
    full_path = os.path.join(delivery_folder, f"{job_name}_{target_lower}_full.txt")
    write_target_only(full_path, job_name, target_language, source_language,
                      rows, generated_at)
    print(f"  wrote {os.path.basename(full_path)}")

    # 3. Human review list
    review_path = os.path.join(delivery_folder, f"{job_name}_chunks_needing_human_review.txt")
    write_review_list(review_path, job_name, target_language, rows, generated_at)
    print(f"  wrote {os.path.basename(review_path)}")

    # 3b. Bible verse review sidecar — only written if the pipeline
    # recorded any content-filter failures during translation.
    bible_review_path = None
    if bible_review:
        bible_review_path = os.path.join(
            delivery_folder, f"{job_name}_bible_verses_needing_verification.txt"
        )
        write_bible_verse_review_file(
            bible_review_path, job_name, target_language,
            bible_review, rows, generated_at,
        )
        print(f"  wrote {os.path.basename(bible_review_path)}")

    # 4. Excel summary
    if make_excel:
        xlsx_path = os.path.join(delivery_folder, f"{job_name}_evaluation_summary.xlsx")
        if write_excel_summary(xlsx_path, job_name, target_language, source_language,
                               rows, generated_at):
            print(f"  wrote {os.path.basename(xlsx_path)}")

    # 5. Small index file so the team knows what's inside
    index_path = os.path.join(delivery_folder, "README.txt")
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(f"{job_name} delivery package\n")
        f.write(f"Generated: {generated_at}\n")
        f.write(f"Target: {target_language} | Source: {source_language}\n")
        f.write(f"Total chunks: {len(rows)}\n")
        f.write(f"Flagged for human review: {flagged_count}\n\n")
        f.write("Files in this folder:\n")
        f.write(f"  {os.path.basename(cbc_path)}\n")
        f.write("    Side-by-side source + translation + scores per chunk.\n")
        f.write(f"  {os.path.basename(full_path)}\n")
        f.write(f"    Complete {target_language} translation only.\n")
        f.write(f"  {os.path.basename(review_path)}\n")
        f.write("    Chunks needing human review, with AI explanations.\n")
        if bible_review_path:
            f.write(f"  {os.path.basename(bible_review_path)}\n")
            f.write("    Bible verse references the API refused to look up.\n")
            f.write("    The translator produced inline renderings for these\n")
            f.write("    verses — a human MUST verify them against the\n")
            f.write(f"    published {target_language} Bible before delivery.\n")
        if make_excel:
            f.write(f"  {job_name}_evaluation_summary.xlsx\n")
            f.write("    Per-chunk scores table with averages.\n")

    print(f"\n  DONE. Delivery folder: {delivery_folder}")
    return delivery_folder


# =====================================================================
# CLI
# =====================================================================

def interactive_pick(batch_dir):
    """Show a numbered menu of jobs and return the one the user picks."""
    jobs = list_job_folders(batch_dir)
    if not jobs:
        raise SystemExit(f"ERROR: no job folders found in {batch_dir}")
    print("\nAvailable jobs:")
    for i, job in enumerate(jobs, start=1):
        job_path = os.path.join(batch_dir, job)
        status = assess_job_readiness(job_path)
        print(f"  [{i:2d}] {job:35s}  {status}")
    print()
    try:
        raw = input("Pick a number (or q to quit): ").strip()
    except EOFError:
        raise SystemExit("No selection made. Exiting.")
    if raw.lower() in ("q", "quit", "exit", ""):
        raise SystemExit("No selection made. Exiting.")
    try:
        idx = int(raw)
    except ValueError:
        raise SystemExit(f"ERROR: '{raw}' is not a number.")
    if not 1 <= idx <= len(jobs):
        raise SystemExit(f"ERROR: {idx} is out of range (1..{len(jobs)}).")
    return jobs[idx - 1]


def cmd_list(batch_dir):
    """Implement --list."""
    jobs = list_job_folders(batch_dir)
    if not jobs:
        print(f"No job folders found in {batch_dir}")
        return
    print(f"Jobs in {batch_dir}:")
    for job in jobs:
        job_path = os.path.join(batch_dir, job)
        status = assess_job_readiness(job_path)
        print(f"  {job:35s}  {status}")


def main():
    parser = argparse.ArgumentParser(
        description="Package one translation job into team deliverables",
    )
    parser.add_argument(
        "--job", type=str, default=None,
        help="Name of the job folder to package (e.g. wauwatosa_spanish). "
             "If omitted, you'll get an interactive picker.",
    )
    parser.add_argument(
        "--list", action="store_true",
        help="List job folders found in the batch directory, with readiness "
             "status, and exit without packaging anything.",
    )
    parser.add_argument(
        "--batch-dir", type=str, default=DEFAULT_BATCH_DIR,
        help=f"Path to the batch folder containing job subfolders "
             f"(default: {DEFAULT_BATCH_DIR})",
    )
    parser.add_argument(
        "--output", type=str, default=DEFAULT_OUTPUT_PARENT,
        help=f"Parent folder where the delivery folder will be created "
             f"(default: {DEFAULT_OUTPUT_PARENT})",
    )
    parser.add_argument(
        "--no-excel", action="store_true",
        help="Skip the Excel summary file (useful if openpyxl is not installed).",
    )
    args = parser.parse_args()

    batch_dir = os.path.abspath(args.batch_dir)
    output_parent = os.path.abspath(args.output)

    if args.list:
        cmd_list(batch_dir)
        return

    if args.job:
        job_name = args.job
    else:
        job_name = interactive_pick(batch_dir)

    job_folder = os.path.join(batch_dir, job_name)
    if not os.path.isdir(job_folder):
        raise SystemExit(f"ERROR: job folder does not exist: {job_folder}")

    os.makedirs(output_parent, exist_ok=True)
    package_job(job_name, job_folder, output_parent, make_excel=not args.no_excel)


if __name__ == "__main__":
    main()
